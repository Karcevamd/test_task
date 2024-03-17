import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import psycopg2
from psycopg2 import sql
from openpyxl import Workbook
from functools import partial

class TableWindow:
    def __init__(self, root, table_name, columns, records):
        self.root = root
        self.root.title(table_name)

        self.table_name = table_name
        self.columns = columns
        self.records = records
        self.filtered_records = records.copy()

        self.frame = ttk.Frame(self.root)
        self.frame.pack(padx=10, pady=10)

        self.tree = ttk.Treeview(self.frame, columns=self.columns, show="headings")
        for col in self.columns:
            self.tree.heading(col, text=col)

        for record in self.records:
            self.tree.insert("", "end", values=record)

        self.tree.grid(row=0, column=0, columnspan=3, sticky="nsew")

        self.column_options = ttk.Combobox(self.frame, values=self.columns)
        self.column_options.set("Выберите столбец")
        self.column_options.grid(row=1, column=0, padx=5, pady=5)

        self.filter_entry = ttk.Entry(self.frame)
        self.filter_entry.grid(row=1, column=1, padx=5, pady=5)

        self.apply_button = ttk.Button(self.frame, text="Применить", command=self.apply_filter)
        self.apply_button.grid(row=2, column=0, columnspan=3, pady=5)

        self.save_button = ttk.Button(self.frame, text="Сохранить в Excel", command=self.save_to_excel)
        self.save_button.grid(row=3, column=0, columnspan=3, pady=5)

    def apply_filter(self):
        keyword = self.filter_entry.get()
        column = self.column_options.get()

        if not keyword:
            self.filtered_records = self.records.copy()
        else:
            if not column:
                messagebox.showerror("Ошибка", "Выберите столбец")
                return

            self.filtered_records = [record for record in self.records if keyword.lower() in str(record[self.columns.index(column)]).lower()]

        self.tree.delete(*self.tree.get_children())

        for record in self.filtered_records:
            self.tree.insert("", "end", values=record)

    def save_to_excel(self):
        wb = Workbook()
        ws = wb.active

        for idx, col in enumerate(self.columns):
            ws.cell(row=1, column=idx+1, value=col)

        records = self.filtered_records if self.filtered_records else self.records

        for row_idx, record in enumerate(records, start=2):
            for col_idx, value in enumerate(record, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(f"{self.table_name}.xlsx")
        messagebox.showinfo("Успех", f"Данные сохранены в {self.table_name}.xlsx")

class DatabaseApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Database App")

        self.conn = psycopg2.connect(dbname="mydatabase", user="postgres", password="0000")
        self.cur = self.conn.cursor()

        self.main_frame = ttk.Frame(self.root)
        self.main_frame.grid(row=0, column=0, sticky="nsew")

        self.table_buttons = ["СУДНО", "МЕСТА_ПОГРУЗКИ", "ГРУЗ", "ПОГРУЗКА"]

        for i, table in enumerate(self.table_buttons):
            button = ttk.Button(self.main_frame, text=table, command=partial(self.open_table_window, table))
            button.grid(row=0, column=i, padx=5, pady=5)

        self.pogruzka_button = ttk.Button(self.main_frame, text="ПОГРУЗКА_ALL", command=self.show_pogruzka_table)
        self.pogruzka_button.grid(row=0, column=len(self.table_buttons), padx=5, pady=5)

    def open_table_window(self, table_name):
        self.cur.execute(sql.SQL("SELECT * FROM {}").format(sql.Identifier(table_name)))
        records = self.cur.fetchall()

        if not records:
            messagebox.showinfo("Информация", "Таблица пуста")
            return

        columns = [desc[0] for desc in self.cur.description]

        table_window = tk.Toplevel(self.root)
        TableWindow(table_window, table_name, columns, records)

    def show_pogruzka_table(self):
        self.cur.execute("""
            SELECT p.НОМЕР_ВЕДОМОСТИ, p.ДАТА, s.НАЗВАНИЕ AS СУДНО, mp.ПОРТ, g.НАЗВАНИЕ AS ГРУЗ, p.КОЛ_ВО, p.СТОИМОСТЬ
            FROM ПОГРУЗКА p
            INNER JOIN СУДНО s ON p.СУДНО = s.ИДЕНТИФИКАТОР
            INNER JOIN МЕСТА_ПОГРУЗКИ mp ON p.МЕСТО_ПОГРУЗКИ = mp.ИДЕНТИФИКАТОР
            INNER JOIN ГРУЗ g ON p.ГРУЗ = g.ИДЕНТИФИКАТОР

        """)
        records = self.cur.fetchall()

        if not records:
            messagebox.showinfo("Информация", "Таблица ПОГРУЗКА пуста")
            return

        columns = [desc[0] for desc in self.cur.description]

        table_window = tk.Toplevel(self.root)
        TableWindow(table_window, "ПОГРУЗКА_ALL", columns, records)

if __name__ == "__main__":
    root = tk.Tk()
    app = DatabaseApp(root)
    root.mainloop()
